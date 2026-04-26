#FUNCIONES
#guardar_excel(datos)
#leer_excel()
# estats()

from openpyxl import load_workbook, Workbook
from time import asctime as tiempo
import os
from statistics import mean, mode, median, stdev, variance

#info de ejemplo
resultados = [{"magnitud": 2,
               "lugar": 4,
               "fecha": 6,
               "profundidad": 8},
              {"magnitud": 3,
               "lugar": 6,
               "fecha": 9,
               "profundidad": 12},
              {"magnitud": 4,
               "lugar": 8,
               "fecha": 12,
               "profundidad": 16},
              {"magnitud": 5,
               "lugar": 10,
               "fecha": 15,
               "profundidad": 20},
              ]

def guardar_excel(resultados):
    #Crea el libro de reportes
    if not os.path.exists("./Reportes.xlsx"):
        libroR = Workbook()
        libroR.save("./Reportes.xlsx")
    #crea un libro
    libro = Workbook()
    arch = "Reporte_del_"+str(tiempo())+".xlsx"
    archi = arch.replace(":", ";")
    archivo = archi.replace(" ", "_")
    libro.save(archivo)

    #Crea una hoja para el reporte
    libro = load_workbook(archivo)
    hoja = libro.active
    hoja.title = "Reporte"
   
    #Encabezado de eventos pendiente

    #roto y meto la información:
    data = list(resultados[1].keys())
    for i in range(0,4):
        tabla = []
        tabla.append(data[i])
        for info in resultados:
            lecturas = list(info.values())
            tabla.append(lecturas[i]) #str(data[i])+"Evento"+str(info)+":"
        hoja.append(tabla)
          
    #guardando libro :D
    libro.save(archivo)

    #Guardando nombre delarchivo
    libroR = load_workbook("./Reportes.xlsx")
    libroR.active.append([archivo])
    libroR.save("./Reportes.xlsx")

def leer_excel():
    #mostramos los reportes existentes
    reportes = load_workbook("./Reportes.xlsx")
    print("Los reportes disponibles son los siguientes:\n")
    for linea in reportes.active.values:
        for valor in linea:
            print(valor)
    #dejamos que el usuario elija un archivo        
    lectura = input("\n Elija un reporte para leer\n:")
    leer = load_workbook(lectura)
    for linea in leer.active.values:
        print(linea,"\n")

def estats():
    #mostramos los reportes existentes
    reportes = load_workbook("./Reportes.xlsx")
    print("Los reportes disponibles son los siguientes:\n")
    for linea in reportes.active.values:
        for valor in linea:
            print(valor)
    #dejamos que el usuario elija un archivo        
    lectura = input("\nElija un reporte para analizar\n:")
    analisis = load_workbook(lectura)

    #creamos el diccionario central
    QuakeInfo = dict()
    for linea in analisis.active.values:
        datos = list()
        for i in linea:
            datos.append(i)
        key = datos[0]
        datos.pop(0)
        QuakeInfo.update({key : datos})

    #UI
    opciones = QuakeInfo.keys()
    print("De los siguientes datos:",)
    for i in opciones:
        print(i)
    x = input("elija cual analizar\n:")
    y = QuakeInfo[x]

    print("De las siguientes operaciones estadisticas")
    print("1. Media")
    print("2. Moda")
    print("3. Mediana")
    print("4. Desviación estandar")
    print("5. varianza")

    op = int(input("Elija numericamente la operación\n:"))

    if op == 1:
        print("La media de",x,"es:",mean(y))
    elif op == 2:
        print("La moda de",x,"es:",mode(y))
    elif op == 3:
        print("La mediana de",x,"es:",median(y))
    elif op == 4:
        print("La desviación estandar de",x,"es:",stdev(y))
    elif op == 5:
        print("La varianza de",x,"es:",variance(y))
    else:
        print("Error por operación inexistente >:v")
    

#Para probar el codigo remover los "#" de las siguientes funciones
#guardar_excel(resultados)
#leer_excel()
#estats()
#print("prueba terminada")


